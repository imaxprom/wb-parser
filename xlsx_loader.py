"""Load articles and queries from xlsx file."""

import re
from openpyxl import load_workbook
import db


def _extract_sku(raw) -> str:
    """Extract numeric WB article from cell value."""
    if isinstance(raw, (int, float)):
        return str(int(raw))
    text = str(raw).strip()
    # If already a clean number — use as is
    if re.fullmatch(r"\d{5,15}", text):
        return text
    # Try to extract from patterns like "Артикул WB: 322000486"
    m = re.search(r"(\d{5,15})", text)
    return m.group(1) if m else ""


def load_from_xlsx(telegram_id: int, file_path: str) -> dict:
    """Load articles and queries from xlsx.

    Expected format:
        Column A: article SKU (numeric WB article ID)
        Column B: search query
    """
    wb = load_workbook(file_path, read_only=True)
    try:
        ws = wb.active
        added = []
        skipped = []

        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue

            sku = _extract_sku(row[0])
            query = str(row[1]).strip()

            if not sku or not query:
                continue

            if sku.lower() in ("артикул", "sku", "article", "id"):
                continue

            article = db.get_article_by_sku(telegram_id, sku)
            if not article:
                article_id = db.add_article(telegram_id, sku)
                if article_id is None:
                    article = db.get_article_by_sku(telegram_id, sku)
                    article_id = article["id"]
            else:
                article_id = article["id"]

            query_id = db.add_query(telegram_id, article_id, query)
            if query_id:
                added.append((sku, query))
            else:
                skipped.append((sku, query, "дубликат"))
    finally:
        wb.close()

    return {"added": added, "skipped": skipped}
